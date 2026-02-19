from django.core.management.base import BaseCommand

import time
import os
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from zoneinfo import ZoneInfo
from openpyxl import load_workbook

from entsoe_api.models import Country, CountryPricePoint


CYPRUS_TZ = ZoneInfo("Asia/Nicosia")


class SeleniumExcelScraper:
    """Automates browser to click Excel export button on TSOC website and download actual market data."""

    def run(self, start_date: str = "today", end_date: str = "today") -> dict:
        """
        Uses Selenium to click export button and download Excel file.

        Args:
            start_date: Start date for data retrieval (YYYY-MM-DD)
            end_date: End date for data retrieval (YYYY-MM-DD)

        Returns:
            dict with status and parsed data entries
        """
        try:
            from selenium.webdriver.common.by import By
            from selenium.webdriver.support.ui import WebDriverWait
            from selenium.webdriver.support import expected_conditions as EC
            import undetected_chromedriver as uc
        except ImportError:
            return {
                "status": "error",
                "error": "Required packages not installed",
                "message": "pip install selenium undetected-chromedriver",
            }

        driver = None
        try:
            chrome_options = uc.ChromeOptions()
            chrome_options.add_argument('--no-sandbox')
            chrome_options.add_argument('--disable-dev-shm-usage')
            chrome_options.add_argument('--window-size=1920,1080')

            download_dir = str(Path.home() / "Downloads" / "tsoc_data")
            os.makedirs(download_dir, exist_ok=True)

            prefs = {
                "download.default_directory": download_dir,
                "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "safebrowsing.enabled": False,
            }
            chrome_options.add_experimental_option("prefs", prefs)

            driver = uc.Chrome(options=chrome_options, use_subprocess=True)

            url = (
                f"https://tsoc.org.cy/competitive-electricity-market/"
                f"dam-volume-prices-graph/?startdt={start_date}&enddt={end_date}"
            )
            driver.get(url)

            time.sleep(15)

            if "just a moment" in driver.title.lower():
                time.sleep(10)

            # Try to find Excel export button
            export_button = None

            try:
                buttons = driver.find_elements(
                    By.XPATH,
                    "//*[contains(translate(text(), 'EXCEL', 'excel'), 'excel')]",
                )
                for btn in buttons:
                    if btn.is_displayed():
                        export_button = btn
                        break
            except Exception:
                pass

            if not export_button:
                selectors = [
                    "//button[contains(@class, 'export')]",
                    "//a[contains(@class, 'export')]",
                    "//button[contains(@id, 'export')]",
                    "//a[contains(@href, 'export')]",
                    "//input[@type='button' and contains(@value, 'Excel')]",
                ]
                for selector in selectors:
                    try:
                        btn = driver.find_element(By.XPATH, selector)
                        if btn.is_displayed():
                            export_button = btn
                            break
                    except Exception:
                        continue

            if not export_button:
                screenshot_path = os.path.join(
                    download_dir,
                    f"page_screenshot_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png",
                )
                driver.save_screenshot(screenshot_path)
                return {
                    "status": "error",
                    "error": "Export button not found",
                    "message": f"Check screenshot: {screenshot_path}",
                    "page_title": driver.title,
                }

            driver.execute_script("arguments[0].click();", export_button)
            time.sleep(10)

            files = sorted(Path(download_dir).glob("*.xlsx"), key=os.path.getmtime, reverse=True)
            if not files:
                files = sorted(Path(download_dir).glob("*.xls"), key=os.path.getmtime, reverse=True)

            if not files:
                return {
                    "status": "error",
                    "error": "File not downloaded",
                    "message": f"No Excel file found in {download_dir}",
                }

            excel_file = files[0]
            data_entries = self._parse_excel(str(excel_file), start_date)

            return {
                "status": "success",
                "source_url": url,
                "start_date": start_date,
                "end_date": end_date,
                "excel_file": str(excel_file),
                "data": data_entries,
                "total_entries": len(data_entries),
            }

        except Exception as e:
            return {
                "status": "error",
                "error_type": type(e).__name__,
                "error_message": str(e),
                "suggestion": "Ensure Chrome is installed.",
            }

        finally:
            if driver:
                driver.quit()

    def _parse_excel(self, excel_file_path: str, date_str: str) -> list[dict]:
        """Parse Excel file and return list of {datetime_utc, price} dicts."""
        base_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        data_entries = []

        try:
            workbook = load_workbook(excel_file_path)
            sheet = workbook.active

            time_col = None
            price_col = None

            for row_idx in range(1, min(11, sheet.max_row + 1)):
                row_values = [str(cell.value).lower() if cell.value else '' for cell in sheet[row_idx]]

                for col_idx, value in enumerate(row_values):
                    if any(kw in value for kw in ['time', 'hour', 'period', 'interval']):
                        time_col = col_idx
                    if any(kw in value for kw in ['price', 'dam', '€/mwh', 'eur']):
                        price_col = col_idx

                if time_col is not None and price_col is not None:
                    for data_row_idx in range(row_idx + 1, sheet.max_row + 1):
                        row = sheet[data_row_idx]
                        time_value = row[time_col].value
                        price_value = row[price_col].value

                        if time_value and price_value is not None:
                            # Build full datetime in Cyprus timezone, then convert to UTC
                            if isinstance(time_value, datetime):
                                local_dt = time_value.replace(tzinfo=CYPRUS_TZ)
                            else:
                                ts = str(time_value).strip()
                                try:
                                    t = datetime.strptime(ts, "%H:%M").time()
                                except ValueError:
                                    t = datetime.strptime(ts, "%H:%M:%S").time()
                                local_dt = datetime.combine(base_date, t, tzinfo=CYPRUS_TZ)

                            dt_utc = local_dt.astimezone(ZoneInfo("UTC"))

                            try:
                                if isinstance(price_value, (int, float)):
                                    price = Decimal(str(price_value))
                                else:
                                    price = Decimal(
                                        str(price_value).replace('€', '').replace(',', '.').strip()
                                    )
                                data_entries.append({
                                    "datetime_utc": dt_utc,
                                    "price": price,
                                })
                            except Exception:
                                continue
                    break

        except Exception as e:
            raise RuntimeError(f"Excel parsing error: {e}") from e

        return data_entries


class Command(BaseCommand):
    help = "Scrape TSOC Cyprus electricity DAM prices and save to CountryPricePoint"

    def add_arguments(self, parser):
        parser.add_argument(
            '--start-date', default='today',
            help='Start date (YYYY-MM-DD or "today")',
        )
        parser.add_argument(
            '--end-date', default='today',
            help='End date (YYYY-MM-DD or "today")',
        )

    def handle(self, *args, **options):
        start_date = options['start_date']
        end_date = options['end_date']

        if start_date == 'today':
            start_date = datetime.now().strftime('%Y-%m-%d')
        if end_date == 'today':
            end_date = datetime.now().strftime('%Y-%m-%d')

        self.stdout.write(f"Scraping TSOC data from {start_date} to {end_date}...")

        scraper = SeleniumExcelScraper()
        result = scraper.run(start_date=start_date, end_date=end_date)

        if result["status"] != "success":
            self.stderr.write(self.style.ERROR(
                f"Scraping failed: {result.get('error', result.get('error_message', 'Unknown'))}"
            ))
            if result.get("message"):
                self.stderr.write(f"  {result['message']}")
            return

        data = result["data"]
        if not data:
            self.stderr.write(self.style.WARNING("No data entries parsed from Excel."))
            return

        country = Country.objects.get(iso_code="CY")

        created_count = 0
        updated_count = 0

        for entry in data:
            _, created = CountryPricePoint.objects.update_or_create(
                country=country,
                contract_type="A01",
                datetime_utc=entry["datetime_utc"],
                defaults={
                    "price": entry["price"],
                    "currency": "EUR",
                    "unit": "MWH",
                    "resolution": "PT60M",
                },
            )
            if created:
                created_count += 1
            else:
                updated_count += 1

        self.stdout.write(self.style.SUCCESS(
            f"Done: {created_count} created, {updated_count} updated "
            f"({len(data)} total entries from {result['excel_file']})"
        ))
        for entry in data[:5]:
            self.stdout.write(
                f"  {entry['datetime_utc']:%Y-%m-%d %H:%M UTC}: {entry['price']} EUR/MWh"
            )
        if len(data) > 5:
            self.stdout.write(f"  ... and {len(data) - 5} more")
